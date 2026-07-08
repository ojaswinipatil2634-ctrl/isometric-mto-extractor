"""
Export generation for a persisted MTO extraction run - CSV, JSON, and
Excel, per Phase 8 scope. No new data is computed here - this only
serializes what's already stored on the run; if something wasn't
detected/extracted, it's simply absent from the export rather than
filled in.
"""
import csv
import io
import json

from openpyxl import Workbook

from app.models.mto_run import MTOExtractionRun


def _run_summary_dict(run: MTOExtractionRun) -> dict:
    return {
        "id": run.id,
        "filename": run.filename,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "drawing_number": run.drawing_number,
        "revision": run.revision,
        "line_number": run.line_number,
        "service": run.service,
        "material_class": run.material_class,
        "nps_values": run.nps_values,
        "extraction_source": run.extraction_source,
        "used_mock": run.used_mock,
        **(run.mto_summary or {}),
        "node_count": run.node_count,
        "edge_count": run.edge_count,
        "branch_count": run.branch_count,
        "dead_end_count": run.dead_end_count,
        "loop_count": run.loop_count,
        "is_fully_connected": run.is_fully_connected,
        "duplicate_fitting_count": run.duplicate_fitting_count,
        "symbol_detection_enabled": run.symbol_detection_enabled,
        "symbol_detection_reason": run.symbol_detection_reason,
        "processing_time_ms": run.processing_time_ms,
        "warnings": run.warnings,
    }


_ITEM_COLUMNS = [
    "item_no", "category", "description", "size_nps", "schedule_rating",
    "material_spec", "end_type", "quantity", "unit", "length_m", "confidence", "remarks",
]


def to_json_bytes(run: MTOExtractionRun) -> bytes:
    payload = _run_summary_dict(run)
    payload["items"] = run.items or []
    payload["hardware"] = run.hardware
    payload["violations"] = run.violations
    return json.dumps(payload, indent=2).encode("utf-8")


def to_csv_bytes(run: MTOExtractionRun) -> bytes:
    """
    A single CSV with labeled sections (summary, MTO items, hardware,
    violations) - simplest format that still opens correctly in any
    spreadsheet tool without needing a zip of multiple files.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["# Summary"])
    for key, value in _run_summary_dict(run).items():
        if isinstance(value, list):
            value = "; ".join(str(v) for v in value)
        writer.writerow([key, value])
    writer.writerow([])

    writer.writerow(["# Material Take-Off"])
    writer.writerow(_ITEM_COLUMNS)
    for item in (run.items or []):
        writer.writerow([item.get(col, "") for col in _ITEM_COLUMNS])
    writer.writerow([])

    writer.writerow(["# Hardware (detection-based, bonus cross-check)"])
    writer.writerow(["item_type", "node_id", "quantity", "size", "is_estimated"])
    for item in run.hardware:
        writer.writerow(
            [item["item_type"], item["node_id"], item["quantity"], item["size"], item["is_estimated"]]
        )
    writer.writerow([])

    writer.writerow(["# Violations"])
    writer.writerow(["rule_code", "severity", "message", "node_ids"])
    for v in run.violations:
        writer.writerow(
            [v["rule_code"], v["severity"], v["message"], ";".join(str(n) for n in v["node_ids"])]
        )

    return buffer.getvalue().encode("utf-8")


def to_xlsx_bytes(run: MTOExtractionRun) -> bytes:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Field", "Value"])
    for key, value in _run_summary_dict(run).items():
        if isinstance(value, list):
            value = "; ".join(str(v) for v in value)
        summary_sheet.append([key, value])

    items_sheet = workbook.create_sheet("MTO Items")
    items_sheet.append(_ITEM_COLUMNS)
    for item in (run.items or []):
        items_sheet.append([item.get(col, "") for col in _ITEM_COLUMNS])

    hardware_sheet = workbook.create_sheet("Hardware")
    hardware_sheet.append(["item_type", "node_id", "quantity", "size", "is_estimated"])
    for item in run.hardware:
        hardware_sheet.append(
            [item["item_type"], item["node_id"], item["quantity"], item["size"], item["is_estimated"]]
        )

    violations_sheet = workbook.create_sheet("Violations")
    violations_sheet.append(["rule_code", "severity", "message", "node_ids"])
    for v in run.violations:
        violations_sheet.append(
            [v["rule_code"], v["severity"], v["message"], ";".join(str(n) for n in v["node_ids"])]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
