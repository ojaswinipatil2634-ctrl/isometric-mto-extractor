import csv
import io
import json
from datetime import datetime, timezone

import openpyxl

from app.models.mto_run import MTOExtractionRun
from app.services.mto.export_service import to_csv_bytes, to_json_bytes, to_xlsx_bytes


def _make_run() -> MTOExtractionRun:
    run = MTOExtractionRun(
        filename="drawing.png",
        drawing_number="MTO-001",
        revision="A",
        line_number='6"-CS-1001-A1A',
        service="COOLING WATER",
        material_class="150",
        nps_values=["6"],
        node_count=4,
        edge_count=3,
        branch_count=1,
        dead_end_count=3,
        loop_count=0,
        is_fully_connected=True,
        hardware=[{"item_type": "gasket", "node_id": 1, "quantity": 1, "size": '6" NPS', "is_estimated": False}],
        violations=[{"rule_code": "MISSING_FITTING", "severity": "warning", "message": "msg", "node_ids": [2]}],
        items=[
            {
                "item_no": 1, "category": "PIPE", "description": "Pipe, Seamless, BE, ASME B36.10",
                "size_nps": '6"', "schedule_rating": "SCH 40", "material_spec": "ASTM A106 Gr.B",
                "end_type": None, "quantity": 12.4, "unit": "M", "length_m": 12.4,
                "confidence": 0.6, "remarks": "",
            }
        ],
        mto_summary={"total_pipe_length_m": 12.4, "fittings": 0, "flanges": 0, "valves": 0, "gaskets": 0, "bolt_sets": 0},
        extraction_source="mock",
        used_mock=True,
        duplicate_fitting_count=0,
        warnings=["a warning"],
        processing_time_ms=12.3,
    )
    run.id = 1
    run.created_at = datetime(2026, 1, 1, tzinfo=timezone.utc)
    return run


def test_to_json_bytes_round_trips():
    payload = json.loads(to_json_bytes(_make_run()))

    assert payload["drawing_number"] == "MTO-001"
    assert payload["hardware"][0]["item_type"] == "gasket"
    assert payload["violations"][0]["rule_code"] == "MISSING_FITTING"


def test_to_csv_bytes_contains_all_sections():
    text = to_csv_bytes(_make_run()).decode("utf-8")
    rows = list(csv.reader(io.StringIO(text)))

    joined = [",".join(r) for r in rows]
    assert any(r.startswith("# Summary") for r in joined)
    assert any(r.startswith("# Material Take-Off") for r in joined)
    assert any(r.startswith("# Hardware") for r in joined)
    assert any(r.startswith("# Violations") for r in joined)
    assert any("PIPE" in r for r in joined)
    assert any("gasket" in r for r in joined)
    assert any("MISSING_FITTING" in r for r in joined)


def test_to_xlsx_bytes_has_four_sheets_with_data():
    workbook = openpyxl.load_workbook(io.BytesIO(to_xlsx_bytes(_make_run())))

    assert workbook.sheetnames == ["Summary", "MTO Items", "Hardware", "Violations"]

    items_sheet = workbook["MTO Items"]
    item_rows = list(items_sheet.iter_rows(values_only=True))
    assert item_rows[1][1] == "PIPE"

    hardware_sheet = workbook["Hardware"]
    rows = list(hardware_sheet.iter_rows(values_only=True))
    assert rows[0] == ("item_type", "node_id", "quantity", "size", "is_estimated")
    assert rows[1][0] == "gasket"

    violations_sheet = workbook["Violations"]
    v_rows = list(violations_sheet.iter_rows(values_only=True))
    assert v_rows[1][0] == "MISSING_FITTING"
